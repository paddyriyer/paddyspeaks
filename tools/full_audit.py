"""Comprehensive SQL + Python audit. Reads:
   • interview/data/questions.json
   • interview/data/question_schemas.json
   • interview/data/enrichments/*.html
   • /tmp/pglite-audit/results.json  (run the PGlite audit first)

Emits a multi-sheet XLSX at interview/data/audit_issues.xlsx with:
   1. Summary               — top-line counts
   2. By severity P0        — questions that fail to run at all
   3. By severity P1        — correctness / completeness gaps
   4. By severity P2        — quality / nice-to-have
   5. SQL questions         — full per-row audit
   6. Python questions      — full per-row audit
   7. By company            — issue density per company
   8. Enrichment coverage   — which questions have / lack the rich treatment
"""
import json, os, re, ast
from collections import defaultdict, Counter

ROOT = '/home/user/paddyspeaks'
QUESTIONS = f'{ROOT}/interview/data/questions.json'
SCHEMAS = f'{ROOT}/interview/data/question_schemas.json'
ENRICHMENTS_DIR = f'{ROOT}/interview/data/enrichments'
PGLITE_RESULTS = '/tmp/pglite-audit/results.json'
OUT_XLSX = f'{ROOT}/interview/data/audit_issues.xlsx'

with open(QUESTIONS) as f: qs = json.load(f)
with open(SCHEMAS) as f: schemas = json.load(f)
enrichment_files = {f[:-5] for f in os.listdir(ENRICHMENTS_DIR) if f.endswith('.html')}

try:
    with open(PGLITE_RESULTS) as f: pglite_results = json.load(f)
    pglite_by_id = {r['id']: r for r in pglite_results}
except FileNotFoundError:
    pglite_by_id = {}

# ─── Python audit helpers ───────────────────────────────────────────
def python_audit(q):
    """Return list of (severity, kind, message) issues for a Python question."""
    issues = []
    sol = q.get('solution', '') or ''
    if not sol.strip():
        issues.append(('P0', 'empty_solution', 'no solution code at all'))
        return issues

    # P0 — syntax error
    try:
        tree = ast.parse(sol)
    except SyntaxError as e:
        issues.append(('P0', 'syntax_error', f'line {e.lineno}: {e.msg}'))
        return issues

    # Find top-level executable statements (not def/class/import/comment)
    runnable_stmts = []
    has_if_main = False
    for node in tree.body:
        if isinstance(node, (ast.Import, ast.ImportFrom)): continue
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)): continue
        # Strings as top-level (module docstrings) don't count as executable
        if isinstance(node, ast.Expr) and isinstance(node.value, ast.Constant) and isinstance(node.value.value, str):
            continue
        # if __name__ == "__main__": ...
        if (isinstance(node, ast.If) and
            isinstance(node.test, ast.Compare) and
            isinstance(node.test.left, ast.Name) and node.test.left.id == '__name__'):
            has_if_main = True
            continue
        runnable_stmts.append(node)

    # P1 — no runnable demo
    if not has_if_main and not runnable_stmts:
        issues.append(('P1', 'no_demo', 'solution defines only functions/classes — Run will produce no output'))

    # Runtime tag sanity
    rt = q.get('runtime') or ''
    sol_lower = sol.lower()
    if rt == '' or rt is None:
        # Should be null only if it's pure Python (no pandas/numpy/spark/flink)
        if 'pandas' in sol_lower or 'import pd' in sol_lower or 'pd.' in sol_lower:
            issues.append(('P1', 'runtime_tag_missing', 'uses pandas but runtime is null'))
    if rt in ('pyspark',) and 'pyspark' not in sol_lower and 'sparksession' not in sol_lower:
        issues.append(('P1', 'runtime_tag_wrong', "tagged pyspark but doesn't import it"))
    if rt == 'third-party':
        # legitimately unrunnable in Pyodide — flag P2 so reader knows
        issues.append(('P2', 'third_party_runtime', 'uses a library not bundled with Pyodide; Run is disabled'))

    # P2 — question text quality
    qt = (q.get('question') or '').strip()
    if not qt:
        issues.append(('P2', 'no_question_text', 'inline `question` field empty'))
    elif len(qt) < 30:
        issues.append(('P2', 'stub_question_text', f'only {len(qt)} chars'))

    # P2 — enrichment coverage
    if q['id'] not in enrichment_files:
        issues.append(('P2', 'no_enrichment', 'no enrichment HTML at interview/data/enrichments/<id>.html'))

    return issues

# ─── SQL audit helpers ──────────────────────────────────────────────
# Tight ghost list: SQL data-type keywords + pure connectors that have no plausible
# semantic as column names. We deliberately DON'T flag partition / date / timestamp /
# count / current / preceding — those are real column names in many production schemas.
SQL_KEYWORDS = {'as', 'in', 'on', 'or', 'and', 'is', 'if', 'do', 'to', 'by',
                'int', 'integer', 'bigint', 'smallint', 'text', 'varchar', 'char',
                'numeric', 'decimal', 'boolean', 'bool', 'real', 'double', 'precision',
                'unique', 'foreign', 'references', 'returning'}

def sql_audit(q):
    issues = []
    sol = q.get('solution', '') or ''
    if not sol.strip():
        issues.append(('P0', 'empty_solution', 'no solution SQL'))

    # PGlite execution status
    r = pglite_by_id.get(q['id'])
    if r:
        kind = r.get('kind', '')
        info = r.get('info', {})
        info_msg = info if isinstance(info, str) else (info.get('message') if isinstance(info, dict) else '')
        if kind == 'ERR':
            issues.append(('P0', 'pglite_err', info_msg[:120] if info_msg else '(unknown error)'))
        elif kind == 'EMPTY':
            issues.append(('P1', 'pglite_empty', 'query runs but returns 0 rows on synthetic data'))
        elif kind == 'DATA-ISSUE':
            issues.append(('P2', 'pglite_data_issue', info_msg[:120] if info_msg else '(ghost / placeholder values)'))
        elif kind == 'SKIP':
            issues.append(('P2', 'pglite_skip', info_msg[:120] if info_msg else '(no schema / no solution)'))
    else:
        if q.get('runtime') == 'non-sqlite':
            issues.append(('P0', 'non_sqlite', 'runtime=non-sqlite; skipped from PGlite audit'))

    # Multi-method blocks (the `--method 1` `--method 2` competing-answers anti-pattern)
    method_blocks = len(re.findall(r'--\s*method\s+\d', sol.lower()))
    if method_blocks >= 2:
        issues.append(('P1', 'multi_method', f'{method_blocks} competing "--method N" blocks; user has to guess'))

    # Ghost columns in schema
    spec = schemas.get(q['id'], [])
    for t in spec:
        for c in t['columns']:
            cl = c.lower()
            if cl in SQL_KEYWORDS:
                issues.append(('P1', 'ghost_column', f'table {t["table"]}: column "{c}" is a SQL keyword'))
            elif re.match(r'^[a-z]$', cl) and len(t['columns']) > 3 and cl not in ('d','t','c','x','y'):
                issues.append(('P1', 'ghost_column_single', f'table {t["table"]}: suspicious single-letter "{c}"'))

    # Inline schema text drift
    inline = q.get('schema') or ''
    if spec and inline:
        expected = ', '.join(f"{t['table']}({', '.join(t['columns'])})" for t in spec)
        if inline != expected:
            issues.append(('P1', 'schema_text_drift', 'inline schema text disagrees with structured spec'))

    # P2 — quality
    qt = (q.get('question') or '').strip()
    if not qt:
        issues.append(('P2', 'no_question_text', 'inline `question` field empty'))
    elif len(qt) < 30:
        issues.append(('P2', 'stub_question_text', f'only {len(qt)} chars'))

    if q['id'] not in enrichment_files:
        issues.append(('P2', 'no_enrichment', 'no enrichment HTML'))

    return issues

# ─── Run audits ─────────────────────────────────────────────────────
print(f"Auditing {len(qs)} questions...")
all_issues = {}
for q in qs:
    lang = q.get('language')
    if lang == 'sql':
        all_issues[q['id']] = (q, 'sql', sql_audit(q))
    elif lang == 'python':
        all_issues[q['id']] = (q, 'python', python_audit(q))

# Summary stats
sql_qs = [v for v in all_issues.values() if v[1] == 'sql']
py_qs = [v for v in all_issues.values() if v[1] == 'python']
sql_with_issues = sum(1 for q, _, iss in sql_qs if iss)
py_with_issues = sum(1 for q, _, iss in py_qs if iss)

severity_counts = Counter()
kind_counts = Counter()
for q, lang, iss in all_issues.values():
    for sev, kind, _ in iss:
        severity_counts[sev] += 1
        kind_counts[(lang, sev, kind)] += 1

# Enrichment coverage
sql_enriched = sum(1 for q, lang, _ in all_issues.values() if lang == 'sql' and q['id'] in enrichment_files)
py_enriched = sum(1 for q, lang, _ in all_issues.values() if lang == 'python' and q['id'] in enrichment_files)

# Company-level
by_company = defaultdict(lambda: {'sql_total': 0, 'sql_issues': 0, 'python_total': 0, 'python_issues': 0,
                                  'sql_enriched': 0, 'python_enriched': 0})
for q, lang, iss in all_issues.values():
    co = q.get('company') or '(unknown)'
    bucket = by_company[co]
    bucket[f'{lang}_total'] += 1
    if iss: bucket[f'{lang}_issues'] += 1
    if q['id'] in enrichment_files: bucket[f'{lang}_enriched'] += 1

print(f"\n=== Audit summary ===")
print(f"SQL questions:    {len(sql_qs)} total, {sql_with_issues} with issues ({100*sql_with_issues/len(sql_qs):.0f}%)")
print(f"  enriched:       {sql_enriched}/{len(sql_qs)} ({100*sql_enriched/len(sql_qs):.0f}%)")
print(f"Python questions: {len(py_qs)} total, {py_with_issues} with issues ({100*py_with_issues/len(py_qs):.0f}%)")
print(f"  enriched:       {py_enriched}/{len(py_qs)} ({100*py_enriched/len(py_qs):.0f}%)")
print(f"\nBy severity:")
for sev in ['P0', 'P1', 'P2']:
    print(f"  {sev}: {severity_counts[sev]}")
print(f"\nTop issue kinds:")
for (lang, sev, kind), cnt in sorted(kind_counts.items(), key=lambda x: -x[1])[:15]:
    print(f"  [{lang:6s}/{sev}] {kind:30s} {cnt}")

# ─── Write XLSX ─────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("\nopenpyxl not installed; cannot write XLSX. Run: pip install openpyxl")
    raise SystemExit(1)

wb = openpyxl.Workbook()
HEADER_FILL = PatternFill(start_color='1A2332', end_color='1A2332', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, name='Helvetica', size=10)
SEV_COLOR = {'P0': 'FFCCCC', 'P1': 'FFE5B4', 'P2': 'FFF8DC'}

def write_header(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical='center')
    ws.freeze_panes = 'A2'

def autosize(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)

# Sheet 1 — Summary
ws = wb.active
ws.title = 'Summary'
ws['A1'] = 'PaddySpeaks · Comprehensive Audit'
ws['A1'].font = Font(bold=True, size=16, name='Helvetica')
ws['A3'] = 'Generated'
ws['B3'] = __import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M UTC')
ws['A5'] = 'Counts'
ws['A5'].font = Font(bold=True)
rows = [
    ['Total questions', len(qs), ''],
    ['  SQL', len(sql_qs), f'{sql_enriched} enriched ({100*sql_enriched/len(sql_qs):.0f}%)'],
    ['  Python', len(py_qs), f'{py_enriched} enriched ({100*py_enriched/len(py_qs):.0f}%)'],
    ['', '', ''],
    ['Total enrichment files', len(enrichment_files), ''],
    ['  Cross-referencing valid questions', sum(1 for f in enrichment_files if any(q['id']==f for q in qs)), ''],
    ['', '', ''],
    ['Questions with at least one issue', sql_with_issues + py_with_issues, f'{100*(sql_with_issues+py_with_issues)/len(qs):.0f}%'],
    ['  SQL', sql_with_issues, f'{100*sql_with_issues/len(sql_qs):.0f}%'],
    ['  Python', py_with_issues, f'{100*py_with_issues/len(py_qs):.0f}%'],
    ['', '', ''],
    ['By severity', '', ''],
    ['  P0 — fail to run', severity_counts['P0'], 'highest priority'],
    ['  P1 — correctness / completeness', severity_counts['P1'], 'next priority'],
    ['  P2 — quality / nice-to-have', severity_counts['P2'], 'long tail'],
]
for r, row in enumerate(rows, start=6):
    for c, val in enumerate(row, start=1):
        ws.cell(row=r, column=c, value=val)
autosize(ws)

# Helper: write a per-question issue sheet
def write_issue_sheet(sheet_name, severity_filter):
    ws = wb.create_sheet(sheet_name)
    headers = ['id', 'language', 'company', 'difficulty', 'title', 'severity', 'kind', 'detail']
    write_header(ws, headers)
    rows = []
    for q, lang, iss in all_issues.values():
        for sev, kind, detail in iss:
            if severity_filter and sev != severity_filter: continue
            rows.append([q['id'], lang, q.get('company') or '', q.get('difficulty') or '',
                         (q.get('title') or '')[:80], sev, kind, detail])
    # Sort: severity (P0 first), then by company, then by id
    sev_order = {'P0': 0, 'P1': 1, 'P2': 2}
    rows.sort(key=lambda r: (sev_order.get(r[5], 9), r[2], r[0]))
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 6 and val in SEV_COLOR:
                cell.fill = PatternFill(start_color=SEV_COLOR[val], end_color=SEV_COLOR[val], fill_type='solid')
    autosize(ws)
    return len(rows)

p0_count = write_issue_sheet('P0 — fail to run', 'P0')
p1_count = write_issue_sheet('P1 — correctness gaps', 'P1')
p2_count = write_issue_sheet('P2 — quality', 'P2')

# Sheet — All SQL questions
ws = wb.create_sheet('SQL questions')
write_header(ws, ['id', 'company', 'difficulty', 'title', 'runtime', 'pglite_kind', 'issue_count',
                   'has_enrichment', 'has_question_text'])
rows = []
for q, lang, iss in all_issues.values():
    if lang != 'sql': continue
    r = pglite_by_id.get(q['id'])
    pkind = r['kind'] if r else ('non-sqlite-skip' if q.get('runtime')=='non-sqlite' else '?')
    rows.append([q['id'], q.get('company') or '', q.get('difficulty') or '',
                 (q.get('title') or '')[:80], q.get('runtime') or '', pkind, len(iss),
                 'yes' if q['id'] in enrichment_files else 'no',
                 'yes' if (q.get('question') or '').strip() else 'no'])
rows.sort(key=lambda r: (-r[6], r[1], r[0]))
for r, row in enumerate(rows, start=2):
    for c, val in enumerate(row, start=1):
        ws.cell(row=r, column=c, value=val)
autosize(ws)

# Sheet — All Python questions
ws = wb.create_sheet('Python questions')
write_header(ws, ['id', 'company', 'difficulty', 'title', 'runtime', 'issue_count',
                   'has_enrichment', 'has_question_text'])
rows = []
for q, lang, iss in all_issues.values():
    if lang != 'python': continue
    rows.append([q['id'], q.get('company') or '', q.get('difficulty') or '',
                 (q.get('title') or '')[:80], q.get('runtime') or '', len(iss),
                 'yes' if q['id'] in enrichment_files else 'no',
                 'yes' if (q.get('question') or '').strip() else 'no'])
rows.sort(key=lambda r: (-r[5], r[1], r[0]))
for r, row in enumerate(rows, start=2):
    for c, val in enumerate(row, start=1):
        ws.cell(row=r, column=c, value=val)
autosize(ws)

# Sheet — by company
ws = wb.create_sheet('By company')
write_header(ws, ['company', 'sql_total', 'sql_with_issues', 'sql_enriched',
                  'py_total', 'py_with_issues', 'py_enriched', 'total', 'enriched_pct'])
rows = []
for co, b in by_company.items():
    total = b['sql_total'] + b['python_total']
    enriched = b['sql_enriched'] + b['python_enriched']
    rows.append([co, b['sql_total'], b['sql_issues'], b['sql_enriched'],
                 b['python_total'], b['python_issues'], b['python_enriched'],
                 total, f'{100*enriched/total:.0f}%' if total else '0%'])
rows.sort(key=lambda r: -r[7])
for r, row in enumerate(rows, start=2):
    for c, val in enumerate(row, start=1):
        ws.cell(row=r, column=c, value=val)
autosize(ws)

# Sheet — enrichment coverage by topic-cluster (heuristic: id prefix)
ws = wb.create_sheet('Enrichment coverage')
write_header(ws, ['id prefix / cluster', 'total questions', 'enriched', 'coverage %', 'first 5 unenriched'])
by_prefix = defaultdict(lambda: {'total': 0, 'enriched': 0, 'unenriched': []})
for q, lang, _ in all_issues.values():
    prefix = q['id'].split('-')[0]
    b = by_prefix[prefix]
    b['total'] += 1
    if q['id'] in enrichment_files:
        b['enriched'] += 1
    elif len(b['unenriched']) < 5:
        b['unenriched'].append(q['id'])
rows = []
for prefix, b in sorted(by_prefix.items(), key=lambda x: -x[1]['total']):
    pct = 100 * b['enriched'] / b['total'] if b['total'] else 0
    rows.append([prefix, b['total'], b['enriched'], f'{pct:.0f}%', ', '.join(b['unenriched'])])
for r, row in enumerate(rows, start=2):
    for c, val in enumerate(row, start=1):
        ws.cell(row=r, column=c, value=val)
autosize(ws)

wb.save(OUT_XLSX)
print(f"\nWrote {OUT_XLSX}")
print(f"Sheets: Summary, P0/P1/P2, SQL questions, Python questions, By company, Enrichment coverage")
